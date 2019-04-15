from openpyxl import *
from tkinter import *
from tkinter.font import Font


# place your .xlsx file in () like:('C:\\Users\\...\\OneDrive\\Desktop\\gameInventory.xlsx') 
wb = load_workbook('') 


sheet = wb.active 


def excel(): 
	
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 21
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 22
	sheet.column_dimensions['E'].width = 15
	sheet.column_dimensions['F'].width = 11
	sheet.column_dimensions['G'].width = 11

	
	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Console/PC"
	sheet.cell(row=1, column=3).value = "Type"
	sheet.cell(row=1, column=4).value = "Release Date"
	sheet.cell(row=1, column=5).value = "Online"
	sheet.cell(row=1, column=6).value = "Rate"
	sheet.cell(row=1, column=7).value = "ESRB"



def focus1(event): 
	console_field.focus_set() 


 
def focus2(event): 
	
	type_field.focus_set() 


 
def focus3(event): 
	
	release_date_field.focus_set() 



def focus4(event): 
	 
	online_field.focus_set() 


 
def focus5(event): 
	
	rate_field.focus_set() 



def focus6(event): 
	
	esrb_field.focus_set() 



def clear(): 
	
	
	name_field.delete(0, END) 
	console_field.delete(0, END) 
	type1_field.delete(0, END) 
	release_date_field.delete(0, END) 
	online_field.delete(0, END) 
	rate_field.delete(0, END) 
	esrb_field.delete(0, END) 



 
def insert(): 
	
	 
	if (name_field.get() == "" and
		console_field.get() == "" and
		type1_field.get() == "" and
		release_date_field.get() == "" and
		online_field.get() == "" and
		rate_field.get() == "" and
		esrb_field.get() == ""): 
			
		print("empty input") 

	else: 

		
		current_row = sheet.max_row 
		current_column = sheet.max_column 

		 
		sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
		sheet.cell(row=current_row + 1, column=2).value = console_field.get() 
		sheet.cell(row=current_row + 1, column=3).value = type1_field.get() 
		sheet.cell(row=current_row + 1, column=4).value = release_date_field.get() 
		sheet.cell(row=current_row + 1, column=5).value = online_field.get() 
		sheet.cell(row=current_row + 1, column=6).value = rate_field.get() 
		sheet.cell(row=current_row + 1, column=7).value = esrb_field.get() 

                # enter This format again # place your .xlsx file in () like:('C:\\Users\\...\\OneDrive\\Desktop\\gameInventory.xlsx')
		wb.save('') 

		 
		name_field.focus_set() 

		
		clear() 


 
if __name__ == "__main__": 
	
	 
	root = Tk() 

	 
	root.configure(background='#111214') 

	
	root.title("Gambrary") 

	 
	root.geometry("500x300") 

	excel() 

	 
	heading = Label(root, text="My Gambrary", bg="#111214", fg="#fff") 

	 
	name = Label(root, text="Name", bg="#111214", foreground="#fff") 

	 
	console = Label(root, text="Console/PC", bg="#111214", foreground="#fff") 

	 
	type1 = Label(root, text="Type", bg="#111214", foreground="#fff") 

	 
	release_date = Label(root, text="Release Date: ", bg="#111214", foreground="#fff") 

	 
	online = Label(root, text="Online Play", bg="#111214", foreground="#fff") 

	 
	rate = Label(root, text="My Rating", bg="#111214", foreground="#fff") 

	 
	esrb = Label(root, text="ESRB", bg="#111214", foreground="#fff") 

	 
	heading.grid(row=0, column=1) 
	name.grid(row=1, column=0) 
	console.grid(row=2, column=0) 
	type1.grid(row=3, column=0) 
	release_date.grid(row=4, column=0) 
	online.grid(row=5, column=0) 
	rate.grid(row=6, column=0) 
	esrb.grid(row=7, column=0) 

	 
	name_field = Entry(root) 
	console_field = Entry(root) 
	type1_field = Entry(root) 
	release_date_field = Entry(root) 
	online_field = Entry(root) 
	rate_field = Entry(root) 
	esrb_field = Entry(root) 

	
	name_field.bind("<Return>", focus1) 

	 
	console_field.bind("<Return>", focus2) 

	 
	type1_field.bind("<Return>", focus3) 

	 
	release_date_field.bind("<Return>", focus4) 

	 
	online_field.bind("<Return>", focus5) 

	
	rate_field.bind("<Return>", focus6) 

	 
	name_field.grid(row=1, column=1, ipadx="100") 
	console_field.grid(row=2, column=1, ipadx="100") 
	type1_field.grid(row=3, column=1, ipadx="100") 
	release_date_field.grid(row=4, column=1, ipadx="100") 
	online_field.grid(row=5, column=1, ipadx="100") 
	rate_field.grid(row=6, column=1, ipadx="100") 
	esrb_field.grid(row=7, column=1, ipadx="100") 

	
	excel() 

	 
	submit = Button(root, text="Submit", fg="#fff", bg="#1b8437", command=insert) 
						 
	submit.grid(row=8, column=1) 

	 
	root.mainloop() 
